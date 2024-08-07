#!/usr/bin/env python3

from argparse import ArgumentParser
import sys
import openpyxl


def read_translations(translation_filename, column):
    translation_workbook = openpyxl.load_workbook(translation_filename)
    ws = translation_workbook.active
    row = 4
    translation_dict = dict()
    while True:
        english = ws["B{0}".format(row)].value
        try:
            translated = ws["{0}{1}".format(column, row)].value.strip().replace("\n", "")
        except AttributeError:
            translated = None
        if not english:
            break
        translation_dict[english] = translated
        row += 1
    return translation_dict


def do_translation(messages, translation_dict, output):
    msgid = ""
    state = None
    for message in messages:
        message = message.strip()
        if message.startswith("msgid"):
            msgid = message.split(" ", maxsplit=1)[1][1:-1]
            state = "MSGID"
        elif message.startswith('"') and state == "MSGID":
            msgid += message[1:-1]
        elif message == 'msgstr ""':
            if msgid not in translation_dict:
                print(message, "not found in translation dictionary")
            try:
                message = 'msgstr "' + translation_dict.get(msgid, "") + '"'
            except TypeError:
                print("TypeError for {}, value is {}".format(msgid, translation_dict.get(msgid, "")))
                sys.exit(1)
        output.write(message + "\n")


def main(translation_filename, column, messages_filename, output_filename):
    translation_dict = read_translations(translation_filename, column.upper())
    with open(messages_filename) as messages:
        with open(output_filename, "w") as output:
            do_translation(messages, translation_dict, output)


if __name__ == "__main__":
    parser = ArgumentParser("Generate a pybabel messages file from a spreadsheet.")
    parser.add_argument(
        "spreadsheet", help="Spreadsheet containing the translations (.xlsx format)"
    )
    parser.add_argument(
        "column", help="Column from the spreadsheet containing the new language (single letter)"
    )
    parser.add_argument("messages", help="messages.pot file from pybabel")
    parser.add_argument(
        "output", help="Destination file e.g. colournaming/translations/en/LC_MESSAGES"
    )
    args = parser.parse_args()

    main(args.spreadsheet, args.column, args.messages, args.output)
